"""Excel file generator for the KEYSTONE simulator.

Produces ``.xlsx`` workbooks whose column headers match the hardcoded
templates in ``app.ingestion.excel_templates`` (``standard_logstat``,
``equipment_readiness``, ``movement_tracker``), ensuring the Excel parser
can correctly identify and extract the data.

Each function returns raw ``bytes`` (from an in-memory ``BytesIO`` buffer)
suitable for direct ``multipart/form-data`` upload to the
``/api/v1/ingestion/excel`` endpoint.
"""

from __future__ import annotations

import io
import random
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if TYPE_CHECKING:
    from simulator.units import UnitState

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center")

# Month abbreviations for DTG formatting (duplicated from mirc.py to avoid
# cross-generator imports — each generator is self-contained).
_MONTH_ABBR = [
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUN",
    "JUL",
    "AUG",
    "SEP",
    "OCT",
    "NOV",
    "DEC",
]


def _format_dtg(dt: datetime) -> str:
    month = _MONTH_ABBR[dt.month - 1]
    return f"{dt.day:02d}{dt.hour:02d}{dt.minute:02d}Z{month}{dt.strftime('%y')}"


def _write_header_block(
    ws,
    title: str,
    unit_name: str,
    dtg: str,
    classification: str = "UNCLASSIFIED // SIM DATA",
) -> int:
    """Write a 4-row header block and return the next available row (1-indexed).

    Row 1: Title (merged across 8 columns)
    Row 2: Unit name
    Row 3: DTG
    Row 4: Classification
    Row 5: blank separator
    """
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell = ws.cell(row=1, column=1, value=f"SIM_{title}")
    cell.font = _TITLE_FONT
    cell.alignment = _CENTER

    ws.cell(row=2, column=1, value=f"UNIT: {unit_name}")
    ws.cell(row=3, column=1, value=f"AS OF: {dtg}")
    ws.cell(row=4, column=1, value=classification)

    return 6  # data starts at row 6 (row 5 is blank separator)


def _style_header_row(ws, row: int, col_count: int) -> None:
    """Apply blue-fill / white-font styling to a header row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _status_from_pct(pct: float) -> str:
    """Derive GREEN / AMBER / RED status from a percentage value."""
    if pct >= 80.0:
        return "GREEN"
    if pct >= 60.0:
        return "AMBER"
    return "RED"


# ---------------------------------------------------------------------------
# LOGSTAT Excel
# ---------------------------------------------------------------------------


def generate_logstat_excel(unit: UnitState, report_time: datetime) -> bytes:
    """Generate a LOGSTAT Excel workbook matching the ``standard_logstat`` template.

    Column headers (row 6): UNIT | CLASS | ITEM | ON HAND | AUTH | DOS | STATUS | REMARKS

    These match the ``required_headers`` in ``excel_templates.LOGSTAT_TEMPLATES``
    for the ``standard_logstat`` template: ``["UNIT", "CLASS", "ITEM", "ON HAND", "AUTH", "DOS"]``.

    Args:
        unit: Unit state to report.
        report_time: Report timestamp.

    Returns:
        Raw bytes of the ``.xlsx`` workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"SIM_LOGSTAT_{unit.abbreviation}".replace("/", "-").replace(" ", "_")

    dtg = _format_dtg(report_time)
    header_row = _write_header_block(ws, "LOGSTAT REPORT", unit.unit_name, dtg)

    # Column headers
    columns = ["UNIT", "CLASS", "ITEM", "ON HAND", "AUTH", "DOS", "STATUS", "REMARKS"]
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name)
    _style_header_row(ws, header_row, len(columns))

    # Data rows
    data_row = header_row + 1
    for item in unit.supply_items:
        pct = round(item.percentage, 1) if item.percentage else 0.0
        status = (
            item.status
            if hasattr(item, "status") and item.status
            else _status_from_pct(pct)
        )
        ws.cell(row=data_row, column=1, value=unit.unit_name)
        ws.cell(row=data_row, column=2, value=item.supply_class)
        ws.cell(row=data_row, column=3, value=item.name)
        ws.cell(row=data_row, column=4, value=int(item.on_hand))
        ws.cell(row=data_row, column=5, value=int(item.required))
        ws.cell(row=data_row, column=6, value=round(item.dos, 1))
        ws.cell(row=data_row, column=7, value=str(status))
        ws.cell(row=data_row, column=8, value="SIM DATA")
        data_row += 1

    # Auto-size columns
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Equipment readiness Excel
# ---------------------------------------------------------------------------


def generate_equipment_excel(unit: UnitState, report_time: datetime) -> bytes:
    """Generate an equipment readiness Excel matching ``equipment_readiness`` template.

    Column headers (row 6): UNIT | TAMCN | NOMENCLATURE | TOTAL | POSS | MC | NMCM | NMCS | MC% | REMARKS

    These match the ``required_headers``: ``["UNIT", "TAMCN", "NOMENCLATURE", "POSS", "MC"]``.

    Args:
        unit: Unit state to report.
        report_time: Report timestamp.

    Returns:
        Raw bytes of the ``.xlsx`` workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"SIM_EQUIP_{unit.abbreviation}".replace("/", "-").replace(" ", "_")

    dtg = _format_dtg(report_time)
    header_row = _write_header_block(ws, "EQUIPMENT READINESS", unit.unit_name, dtg)

    columns = [
        "UNIT",
        "TAMCN",
        "NOMENCLATURE",
        "TOTAL",
        "POSS",
        "MC",
        "NMCM",
        "NMCS",
        "MC%",
        "REMARKS",
    ]
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name)
    _style_header_row(ws, header_row, len(columns))

    data_row = header_row + 1
    for cat in unit.equipment_categories:
        ws.cell(row=data_row, column=1, value=unit.unit_name)
        ws.cell(row=data_row, column=2, value=cat.tamcn)
        ws.cell(row=data_row, column=3, value=cat.nomenclature)
        ws.cell(row=data_row, column=4, value=cat.total_possessed)
        ws.cell(row=data_row, column=5, value=cat.total_possessed)
        ws.cell(row=data_row, column=6, value=cat.mission_capable)
        ws.cell(row=data_row, column=7, value=cat.nmcm)
        ws.cell(row=data_row, column=8, value=cat.nmcs)
        ws.cell(row=data_row, column=9, value=round(cat.readiness_pct, 1))
        ws.cell(row=data_row, column=10, value="SIM DATA")
        data_row += 1

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Convoy manifest Excel
# ---------------------------------------------------------------------------

# Realistic bumper numbers and cargo descriptions
_BUMPER_PREFIXES = ["1M", "2M", "3M", "HQ"]
_CARGO_TYPES = [
    "CL I - MRE/WATER",
    "CL III - JP-8/DIESEL",
    "CL V - AMMUNITION",
    "CL IX - REPAIR PARTS",
    "MIXED CARGO",
    "CL VIII - MEDICAL SUPPLIES",
    "EMPTY (RETURN)",
]
_VEHICLE_TYPES = ["MTVR", "LVS", "HMMWV", "M1083", "LVSR"]


def generate_convoy_manifest_excel(
    convoy_data: dict,
    report_time: datetime,
) -> bytes:
    """Generate a convoy manifest Excel matching ``movement_tracker`` template.

    Column headers (row 6): CONVOY | VEHICLE | BUMPER | CARGO | WEIGHT | ORIGIN | DESTINATION | ETD | ETA

    These match the ``required_headers``: ``["CONVOY", "ORIGIN", "DESTINATION"]``.

    Args:
        convoy_data: Dict with ``convoy_id``, ``origin``, ``destination``,
            ``vehicle_count``, ``eta`` (HHMMZ), and optionally ``vehicles``
            (list of per-vehicle dicts).
        report_time: Report timestamp.

    Returns:
        Raw bytes of the ``.xlsx`` workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SIM_CONVOY_MANIFEST".replace("/", "-").replace(" ", "_")

    dtg = _format_dtg(report_time)
    cid = convoy_data.get("convoy_id", "C-2026-001")
    header_row = _write_header_block(ws, f"CONVOY MANIFEST {cid}", cid, dtg)

    columns = [
        "CONVOY",
        "VEHICLE",
        "BUMPER",
        "CARGO",
        "WEIGHT",
        "ORIGIN",
        "DESTINATION",
        "ETD",
        "ETA",
    ]
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name)
    _style_header_row(ws, header_row, len(columns))

    origin = convoy_data.get("origin", "CSS AREA")
    dest = convoy_data.get("destination", "CAMP WILSON")
    eta = convoy_data.get("eta", "1200Z")
    veh_count = convoy_data.get("vehicle_count", 8)
    vehicles = convoy_data.get("vehicles", [])

    # Generate per-vehicle rows
    data_row = header_row + 1
    for i in range(veh_count):
        if i < len(vehicles):
            veh = vehicles[i]
        else:
            veh = {
                "type": random.choice(_VEHICLE_TYPES),
                "bumper": f"{random.choice(_BUMPER_PREFIXES)}-{random.randint(100, 999)}",
                "cargo": random.choice(_CARGO_TYPES),
                "weight_lbs": random.randint(5000, 35000),
            }

        # ETD is report_time formatted
        etd_str = report_time.strftime("%d%H%MZ%b%y").upper()

        ws.cell(row=data_row, column=1, value=cid)
        ws.cell(row=data_row, column=2, value=veh.get("type", "MTVR"))
        ws.cell(row=data_row, column=3, value=veh.get("bumper", f"SIM-{i + 1:03d}"))
        ws.cell(row=data_row, column=4, value=veh.get("cargo", "MIXED CARGO"))
        ws.cell(row=data_row, column=5, value=veh.get("weight_lbs", 15000))
        ws.cell(row=data_row, column=6, value=origin)
        ws.cell(row=data_row, column=7, value=dest)
        ws.cell(row=data_row, column=8, value=etd_str)
        ws.cell(row=data_row, column=9, value=eta)
        data_row += 1

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
