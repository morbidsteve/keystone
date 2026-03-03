"""Schema mapping endpoints — canonical fields, data templates, preview, auto-detect."""

import io
import logging
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Query, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth import get_current_user
from app.core.exceptions import BadRequestError, ForbiddenError, NotFoundError
from app.database import get_db
from app.models.canonical_schema import CanonicalField
from app.models.data_template import DataTemplate
from app.models.user import Role, User
from app.schemas.data_template import (
    AutoDetectRequest,
    AutoDetectResponse,
    CanonicalFieldGrouped,
    CanonicalFieldResponse,
    DataTemplateCreate,
    DataTemplateListItem,
    DataTemplateResponse,
    DataTemplateUpdate,
    MappingPreviewRequest,
    MappingPreviewResponse,
    MappingPreviewRow,
)

logger = logging.getLogger(__name__)
router = APIRouter()


# ---------------------------------------------------------------------------
# Canonical Fields
# ---------------------------------------------------------------------------


@router.get("/canonical-fields", response_model=List[CanonicalFieldGrouped])
async def list_canonical_fields(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all canonical fields grouped by entity."""
    result = await db.execute(
        select(CanonicalField).order_by(
            CanonicalField.entity_group,
            CanonicalField.entity_name,
            CanonicalField.field_name,
        )
    )
    fields = result.scalars().all()

    # Group by entity_name
    groups: Dict[str, Dict[str, Any]] = {}
    for f in fields:
        key = str(f.entity_name)
        if key not in groups:
            groups[key] = {
                "entity_name": f.entity_name,
                "entity_group": f.entity_group,
                "fields": [],
            }
        groups[key]["fields"].append(CanonicalFieldResponse.model_validate(f))

    return list(groups.values())


# ---------------------------------------------------------------------------
# Templates CRUD
# ---------------------------------------------------------------------------


@router.get("/templates", response_model=List[DataTemplateListItem])
async def list_templates(
    source_type: Optional[str] = Query(None),
    active_only: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all saved data templates."""
    query = select(DataTemplate)
    if active_only:
        query = query.where(DataTemplate.is_active == True)  # noqa: E712
    if source_type:
        query = query.where(DataTemplate.source_type == source_type)
    query = query.order_by(DataTemplate.updated_at.desc())

    result = await db.execute(query)
    templates = result.scalars().all()

    return [
        DataTemplateListItem(
            id=t.id,
            name=t.name,
            description=t.description,
            source_type=t.source_type,
            field_count=len(t.field_mappings) if t.field_mappings else 0,
            version=t.version,
            is_active=t.is_active,
            created_by=t.created_by,
            created_at=t.created_at,
        )
        for t in templates
    ]


@router.get("/templates/{template_id}", response_model=DataTemplateResponse)
async def get_template(
    template_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get a single template by ID."""
    result = await db.execute(
        select(DataTemplate).where(DataTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise NotFoundError("DataTemplate", template_id)
    return template


@router.post("/templates", response_model=DataTemplateResponse, status_code=201)
async def create_template(
    payload: DataTemplateCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new data template."""
    if current_user.role not in (
        Role.ADMIN,
        Role.OPERATOR,
        Role.S4,
        Role.S3,
        Role.COMMANDER,
    ):
        raise ForbiddenError(
            "Only admin/operator/S4/S3/commander roles can create templates"
        )

    template = DataTemplate(
        name=payload.name,
        description=payload.description,
        source_type=payload.source_type,
        field_mappings=payload.field_mappings,
        header_patterns=payload.header_patterns,
        version=1,
        is_active=True,
        created_by=current_user.id,
    )
    db.add(template)
    await db.flush()
    await db.refresh(template)
    return template


@router.put("/templates/{template_id}", response_model=DataTemplateResponse)
async def update_template(
    template_id: int,
    payload: DataTemplateUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update an existing template (increments version)."""
    if current_user.role not in (
        Role.ADMIN,
        Role.OPERATOR,
        Role.S4,
        Role.S3,
        Role.COMMANDER,
    ):
        raise ForbiddenError(
            "Only admin/operator/S4/S3/commander roles can update templates"
        )

    result = await db.execute(
        select(DataTemplate).where(DataTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise NotFoundError("DataTemplate", template_id)

    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(template, key, value)

    template.version = (template.version or 1) + 1
    await db.flush()
    await db.refresh(template)
    return template


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft-delete (deactivate) a template."""
    if current_user.role not in (
        Role.ADMIN,
        Role.OPERATOR,
        Role.S4,
        Role.S3,
        Role.COMMANDER,
    ):
        raise ForbiddenError(
            "Only admin/operator/S4/S3/commander roles can delete templates"
        )

    result = await db.execute(
        select(DataTemplate).where(DataTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise NotFoundError("DataTemplate", template_id)

    template.is_active = False
    await db.flush()

    return {"id": template_id, "message": "Template deactivated"}


# ---------------------------------------------------------------------------
# Preview Mapping
# ---------------------------------------------------------------------------


def _apply_transform(
    value: Any, transform: Optional[str], params: Optional[Dict] = None
) -> Any:
    """Apply a transform to a single value."""
    if value is None:
        return None
    if not transform:
        return value

    try:
        if transform == "string":
            return str(value)
        elif transform == "integer":
            return int(float(str(value).replace(",", "")))
        elif transform == "float":
            return float(str(value).replace(",", ""))
        elif transform == "datetime":
            from dateutil import parser as dt_parser  # type: ignore[import-untyped]

            fmt = params.get("format") if params else None
            if fmt:
                from datetime import datetime

                return datetime.strptime(str(value), fmt).isoformat()
            return dt_parser.parse(str(value)).isoformat()
        elif transform == "regex":
            import re

            pattern = params.get("pattern", "(.*)") if params else "(.*)"
            match = re.search(pattern, str(value))
            if match:
                return match.group(1) if match.lastindex else match.group(0)
            return str(value)
        elif transform == "enum":
            allowed = params.get("values", []) if params else []
            str_val = str(value).upper().strip()
            if allowed and str_val not in [v.upper() for v in allowed]:
                return str_val  # Return anyway, let validation catch it
            return str_val
        else:
            return value
    except Exception:
        return value


@router.post("/preview", response_model=MappingPreviewResponse)
async def preview_mapping(
    payload: MappingPreviewRequest,
    current_user: User = Depends(get_current_user),
):
    """Preview how data would be mapped using the provided field mappings."""
    rows: List[MappingPreviewRow] = []
    error_count = 0
    successful = 0

    for row_data in payload.sample_data[:20]:  # Cap at 20 rows for preview
        mapped: Dict[str, Any] = {}
        errors: List[str] = []

        for source_col, mapping_config in payload.field_mappings.items():
            target_entity = mapping_config.get("target_entity", "")
            target_field = mapping_config.get("target_field", "")
            transform = mapping_config.get("transform")
            transform_params = mapping_config.get("transform_params")

            raw_value = row_data.get(source_col)
            if raw_value is None:
                continue

            try:
                transformed = _apply_transform(raw_value, transform, transform_params)
                mapped_key = (
                    f"{target_entity}.{target_field}" if target_entity else target_field
                )
                mapped[mapped_key] = transformed
            except Exception as e:
                errors.append(f"Transform error on '{source_col}': {str(e)}")

        if errors:
            error_count += len(errors)
        else:
            successful += 1

        rows.append(MappingPreviewRow(source=row_data, mapped=mapped, errors=errors))

    return MappingPreviewResponse(
        rows=rows,
        total_rows=len(rows),
        successful_rows=successful,
        error_count=error_count,
    )


# ---------------------------------------------------------------------------
# Auto-Detect Template
# ---------------------------------------------------------------------------


@router.post("/auto-detect", response_model=AutoDetectResponse)
async def auto_detect_template(
    payload: AutoDetectRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Given file headers, try to auto-detect a matching saved template."""
    normalized_headers = [h.strip().upper() for h in payload.headers if h]

    result = await db.execute(
        select(DataTemplate).where(
            DataTemplate.is_active == True  # noqa: E712
        )
    )
    templates = result.scalars().all()

    best_template = None
    best_score = 0.0

    for template in templates:
        if not template.header_patterns:
            continue

        pattern_headers = [
            p.strip().upper() for p in list(template.header_patterns) if p
        ]
        if not pattern_headers:
            continue

        # Count how many pattern headers appear in the file headers
        matches = sum(
            1
            for pattern in pattern_headers
            if any(pattern in header for header in normalized_headers)
        )
        score = matches / len(pattern_headers) if pattern_headers else 0

        if score > best_score and score >= 0.4:
            best_score = score
            best_template = template

    if best_template and best_score >= 0.4:
        return AutoDetectResponse(
            matched=True,
            template=DataTemplateResponse.model_validate(best_template),
            confidence=round(best_score, 2),
            message=f"Matched template '{best_template.name}' with {round(best_score * 100)}% confidence",
        )

    return AutoDetectResponse(
        matched=False,
        template=None,
        confidence=0.0,
        message="No matching template found. Use the Schema Mapper to create a new mapping.",
    )


# ---------------------------------------------------------------------------
# Template Test
# ---------------------------------------------------------------------------


@router.post("/templates/{template_id}/test", response_model=MappingPreviewResponse)
async def test_template(
    template_id: int,
    sample_data: List[Dict[str, Any]],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Test a saved template against sample data."""
    result = await db.execute(
        select(DataTemplate).where(DataTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise NotFoundError("DataTemplate", template_id)

    # Reuse preview logic
    preview_request = MappingPreviewRequest(
        field_mappings=template.field_mappings,
        sample_data=sample_data,
    )
    return await preview_mapping(preview_request, current_user)


# ---------------------------------------------------------------------------
# File Upload for Preview (multipart)
# ---------------------------------------------------------------------------


@router.post("/upload-preview")
async def upload_file_for_preview(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Upload a file and return its headers + first rows for mapping UI."""
    if not file.filename:
        raise BadRequestError("No file provided")

    content = await file.read()
    filename = file.filename.lower()

    headers: List[str] = []
    sample_rows: List[Dict[str, Any]] = []

    if filename.endswith((".xlsx", ".xls")):
        headers, sample_rows = _parse_excel_preview(content)
    elif filename.endswith(".csv"):
        headers, sample_rows = _parse_csv_preview(content)
    else:
        raise BadRequestError("Unsupported file type. Accepted: .xlsx, .xls, .csv")

    return {
        "file_name": file.filename,
        "headers": headers,
        "sample_rows": sample_rows,
        "row_count": len(sample_rows),
    }


def _parse_excel_preview(content: bytes) -> tuple:
    """Parse Excel file and return headers + sample rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise BadRequestError("openpyxl not installed on the server")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 1:
        return [], []

    # Find header row
    header_row_idx = 0
    headers = []
    for idx, row in enumerate(rows):
        non_empty = [str(cell).strip() if cell else "" for cell in row]
        if sum(1 for c in non_empty if c) >= 2:
            header_row_idx = idx
            headers = non_empty
            break

    if not headers:
        return [], []

    # Build sample rows (up to 10)
    sample_rows = []
    for row in rows[header_row_idx + 1 : header_row_idx + 11]:
        row_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = cell
        if any(v is not None for v in row_dict.values()):
            sample_rows.append(row_dict)

    # Filter out empty header columns
    headers = [h for h in headers if h]

    return headers, sample_rows


def _parse_csv_preview(content: bytes) -> tuple:
    """Parse CSV file and return headers + sample rows."""
    import csv

    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))

    rows = list(reader)
    if len(rows) < 1:
        return [], []

    headers = [h.strip() for h in rows[0] if h.strip()]

    sample_rows = []
    for row in rows[1:11]:
        row_dict = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = val.strip() if val else None
        if any(v is not None for v in row_dict.values()):
            sample_rows.append(row_dict)

    return headers, sample_rows
