"""Parse Excel files into structured logistics records."""

import io
import logging
from typing import Any, Dict, List, Optional, Tuple

from app.ingestion.excel_templates import identify_template, map_columns

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Database-backed template matching
# ---------------------------------------------------------------------------


def _match_db_template(headers: List[str], db_templates: List[Any]) -> Optional[Tuple[Any, float]]:
    """Try to match headers against DataTemplate records from the database.

    Args:
        headers: Normalized header strings from the file.
        db_templates: List of DataTemplate model instances.

    Returns:
        (template, confidence) or None.
    """
    normalized = [h.strip().upper() for h in headers if h]
    best = None
    best_score = 0.0

    for template in db_templates:
        if not template.header_patterns:
            continue
        patterns = [p.strip().upper() for p in template.header_patterns if p]
        if not patterns:
            continue

        matches = sum(
            1 for p in patterns if any(p in hdr for hdr in normalized)
        )
        score = matches / len(patterns)
        if score > best_score and score >= 0.4:
            best_score = score
            best = template

    if best:
        return best, best_score
    return None


def _apply_db_template(row_data: Dict[str, Any], field_mappings: Dict[str, Dict]) -> Dict[str, Any]:
    """Apply a DataTemplate's field_mappings to a single row dict.

    field_mappings format:
    {
        "SOURCE COL": {
            "target_entity": "supply_status",
            "target_field": "on_hand_qty",
            "transform": "float"
        },
        ...
    }
    """
    mapped: Dict[str, Any] = {}
    for source_col, mapping in field_mappings.items():
        # Try exact match first, then case-insensitive
        raw_value = row_data.get(source_col)
        if raw_value is None:
            for k, v in row_data.items():
                if k.strip().upper() == source_col.strip().upper():
                    raw_value = v
                    break

        if raw_value is None:
            continue

        target_field = mapping.get("target_field", source_col)
        transform = mapping.get("transform")

        try:
            if transform == "float":
                raw_value = float(str(raw_value).replace(",", ""))
            elif transform == "integer":
                raw_value = int(float(str(raw_value).replace(",", "")))
            elif transform == "string":
                raw_value = str(raw_value)
        except (ValueError, TypeError):
            pass

        mapped[target_field] = raw_value

    return mapped


def _entity_to_record_type(entity: str) -> str:
    """Map a canonical entity name to a record type string."""
    mapping = {
        "supply_status": "SUPPLY",
        "equipment_status": "EQUIPMENT",
        "movement": "TRANSPORTATION",
    }
    return mapping.get(entity, "UNKNOWN")


def _detect_entity_from_template(field_mappings: Dict) -> str:
    """Detect the primary entity from a template's field mappings."""
    entity_counts: Dict[str, int] = {}
    for _, cfg in field_mappings.items():
        entity = cfg.get("target_entity", "")
        if entity:
            entity_counts[entity] = entity_counts.get(entity, 0) + 1
    if entity_counts:
        return max(entity_counts, key=entity_counts.get)  # type: ignore[arg-type]
    return ""


# ---------------------------------------------------------------------------
# Main parse function
# ---------------------------------------------------------------------------


def parse_excel(
    file_path_or_bytes,
    file_name: str = "",
    db_templates: Optional[List[Any]] = None,
) -> List[Dict]:
    """Parse an Excel file into a list of structured records.

    Accepts either a file path (str) or bytes content.
    Detects the template by:
      1. Checking DataTemplate records from the database (if provided).
      2. Falling back to hardcoded LOGSTAT_TEMPLATES.
      3. Returning raw UNKNOWN records if nothing matches.

    Args:
        file_path_or_bytes: Path string or bytes of the workbook.
        file_name: Original filename for metadata.
        db_templates: Optional list of DataTemplate model instances to check first.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [{"error": "openpyxl not installed", "confidence": 0.0}]

    try:
        if isinstance(file_path_or_bytes, bytes):
            wb = load_workbook(io.BytesIO(file_path_or_bytes), read_only=True, data_only=True)
        else:
            wb = load_workbook(file_path_or_bytes, read_only=True, data_only=True)
    except Exception as e:
        return [{"error": f"Failed to load workbook: {str(e)}", "confidence": 0.0}]

    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # Find header row (first row with text content)
        header_row_idx = None
        headers = []
        for idx, row in enumerate(rows):
            non_empty = [str(cell).strip() if cell else "" for cell in row]
            if sum(1 for c in non_empty if c) >= 3:
                header_row_idx = idx
                headers = non_empty
                break

        if header_row_idx is None:
            continue

        # -----------------------------------------------------------
        # Step 1: Try database-backed DataTemplate matching
        # -----------------------------------------------------------
        db_match = None
        if db_templates:
            db_match = _match_db_template(headers, db_templates)

        if db_match:
            template_obj, confidence = db_match
            entity = _detect_entity_from_template(template_obj.field_mappings)
            record_type = _entity_to_record_type(entity)

            for row in rows[header_row_idx + 1:]:
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_data[headers[col_idx]] = cell
                if not any(v is not None and v != "" for v in row_data.values()):
                    continue

                mapped_data = _apply_db_template(row_data, template_obj.field_mappings)
                records.append({
                    "type": record_type,
                    "data": mapped_data,
                    "confidence": round(confidence, 2),
                    "template": template_obj.name,
                    "template_id": template_obj.id,
                    "source_sheet": sheet_name,
                    "file_name": file_name,
                })

            logger.info(
                f"Matched DB template '{template_obj.name}' "
                f"(id={template_obj.id}) for sheet '{sheet_name}' "
                f"with confidence {confidence:.0%}"
            )
            continue

        # -----------------------------------------------------------
        # Step 2: Fall back to hardcoded template matching
        # -----------------------------------------------------------
        template_match = identify_template(headers)
        if not template_match:
            # Unknown template — extract raw data
            for row in rows[header_row_idx + 1:]:
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_data[headers[col_idx]] = cell
                if any(v is not None for v in row_data.values()):
                    records.append({
                        "type": "UNKNOWN",
                        "data": row_data,
                        "confidence": 0.3,
                        "source_sheet": sheet_name,
                        "file_name": file_name,
                    })
            continue

        template_name, template_config = template_match
        col_mapping = map_columns(headers, template_config)

        # Extract data rows
        for row in rows[header_row_idx + 1:]:
            row_data = {}
            for col_idx, field_name in col_mapping.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        row_data[field_name] = value

            # Skip empty rows
            if not any(v is not None and v != "" for v in row_data.values()):
                continue

            # Determine record type from template
            if template_name == "standard_logstat":
                record_type = "SUPPLY"
            elif template_name == "equipment_readiness":
                record_type = "EQUIPMENT"
            elif template_name == "movement_tracker":
                record_type = "TRANSPORTATION"
            else:
                record_type = "UNKNOWN"

            records.append({
                "type": record_type,
                "data": row_data,
                "confidence": 0.9,
                "template": template_name,
                "source_sheet": sheet_name,
                "file_name": file_name,
            })

    wb.close()
    return records
